"""
QualityPulse — Export Utilities
Helper functions to export pandas DataFrames to Excel with professional styling
and conditional formatting.
"""

import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def format_excel_sheet(worksheet, auto_filter: bool = True, col_widths: dict = None):
    """
    Apply professional corporate styling to an openpyxl worksheet.

    Args:
        worksheet:   openpyxl Worksheet object.
        auto_filter: Whether to enable column auto-filters.
        col_widths:  Optional dict of column letter -> fixed width, e.g. {"A": 35}.
    """
    header_font      = Font(bold=True, color="FFFFFF", size=10)
    header_fill      = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    accent_fill      = PatternFill(start_color="4F8EF7", end_color="4F8EF7", fill_type="solid")  # accent stripe

    border_side  = Side(style="thin", color="CBD5E1")
    thin_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    alt_fill   = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF",  end_color="FFFFFF",  fill_type="solid")

    if worksheet.max_row > 1:
        # -- Header row
        for cell in worksheet[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_alignment
            cell.border    = thin_border

        worksheet.freeze_panes = "A2"
        worksheet.row_dimensions[1].height = 24

        if auto_filter and worksheet.dimensions:
            try:
                worksheet.auto_filter.ref = worksheet.dimensions
            except Exception:
                pass

    # -- Data rows + auto column widths
    col_max: dict[int, int] = {}
    for col_idx, col in enumerate(worksheet.columns, 1):
        for row_idx, cell in enumerate(col, 1):
            if row_idx > 1:
                cell.fill   = alt_fill if row_idx % 2 == 0 else white_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

            # Track max content length for auto-sizing
            try:
                val = cell.value
                txt = str(val) if val is not None else ""
                length = max(len(line) for line in txt.split("\n")) if txt else 0
                col_max[col_idx] = max(col_max.get(col_idx, 0), length)
            except Exception:
                pass

    # Apply column widths
    for col_idx, max_len in col_max.items():
        letter = get_column_letter(col_idx)
        if col_widths and letter in col_widths:
            worksheet.column_dimensions[letter].width = col_widths[letter]
        else:
            # Min width 12, Max width 60
            width = min(max(max_len + 4, 12), 60)
            worksheet.column_dimensions[letter].width = width


def apply_conditional_scrap(worksheet, col_index: int, scrap_target: float):
    """
    Color-code a numeric column by scrap rate thresholds:
      green  = ≤ scrap_target
      amber  = scrap_target < value ≤ scrap_target × 1.75
      red    = > scrap_target × 1.75
    col_index is 1-based.
    """
    warn = scrap_target * 1.75
    green_fill = PatternFill("solid", fgColor="D1FAE5")
    amber_fill = PatternFill("solid", fgColor="FEF3C7")
    red_fill   = PatternFill("solid", fgColor="FEE2E2")

    for row in worksheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
        for cell in row:
            try:
                val = float(cell.value or 0)
                if val <= scrap_target:
                    cell.fill = green_fill
                elif val <= warn:
                    cell.fill = amber_fill
                else:
                    cell.fill = red_fill
            except (TypeError, ValueError):
                pass


def to_excel(df: pd.DataFrame, sheet_name: str = "Veriler") -> bytes:
    """
    Convert a single DataFrame to a styled Excel file in memory.
    Returns raw bytes for download.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        format_excel_sheet(writer.sheets[sheet_name])
    return output.getvalue()
